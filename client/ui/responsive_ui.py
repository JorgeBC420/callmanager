"""
Interfaz Responsive para Tablets y Móviles - CallManager
Diseño adaptativo para Android, iOS y tablets

Autor: CallManager System
Versión: 1.0
"""

import customtkinter as ctk
from tkinter import messagebox, ttk
import logging
from typing import Optional, Callable, Dict, List
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import threading

logger = logging.getLogger(__name__)


class ResponsiveFrame(ctk.CTkFrame):
    """Frame responsivo que se adapta al tamaño de pantalla"""
    
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.screen_width = master.winfo_width()
        self.screen_height = master.winfo_height()
        self.is_mobile = self.screen_width < 768
        self.is_tablet = 768 <= self.screen_width < 1024
        self.is_desktop = self.screen_width >= 1024
        
        self.bind('<Configure>', self._on_resize)
    
    def _on_resize(self, event):
        """Detectar cambios de tamaño"""
        old_mobile = self.is_mobile
        self.screen_width = self.winfo_width()
        self.screen_height = self.winfo_height()
        
        self.is_mobile = self.screen_width < 768
        self.is_tablet = 768 <= self.screen_width < 1024
        self.is_desktop = self.screen_width >= 1024
        
        # Si cambió el modo, relayoutear
        if old_mobile != self.is_mobile:
            self._on_screen_mode_changed()
    
    def _on_screen_mode_changed(self):
        """Override en subclases para relayout responsivo"""
        pass


class ContactEditorWidget(ResponsiveFrame):
    """Widget de edición inline de contactos"""
    
    def __init__(self, master, contact_data: Dict = None, on_save: Callable = None, **kwargs):
        super().__init__(master, **kwargs)
        
        self.contact_data = contact_data or {}
        self.on_save = on_save
        self.is_editing = False
        self.original_name = self.contact_data.get('name', '')
        
        self._create_widgets()
        self._update_layout()
    
    def _create_widgets(self):
        """Crear widgets de edición"""
        # Frame principal
        main_frame = ctk.CTkFrame(self)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # --- Fila 1: Nombre (editable) ---
        name_frame = ctk.CTkFrame(main_frame)
        name_frame.pack(fill='x', pady=5)
        
        ctk.CTkLabel(name_frame, text="Nombre:", font=("Arial", 12, "bold")).pack(
            side='left', padx=5
        )
        
        self.name_entry = ctk.CTkEntry(
            name_frame,
            placeholder_text="Nombre del contacto",
            font=("Arial", 11),
            height=32
        )
        self.name_entry.pack(side='left', fill='both', expand=True, padx=5)
        self.name_entry.insert(0, self.original_name)
        
        self.edit_btn = ctk.CTkButton(
            name_frame,
            text="✏️ Editar",
            font=("Arial", 11),
            width=100,
            command=self._toggle_edit_mode
        )
        self.edit_btn.pack(side='left', padx=5)
        
        self.name_entry.configure(state='disabled')
        
        # --- Fila 2: Estado (desplegable) + Teléfono ---
        contact_frame = ctk.CTkFrame(main_frame)
        contact_frame.pack(fill='x', pady=5)
        
        ctk.CTkLabel(contact_frame, text="Estado:", font=("Arial", 10)).pack(
            side='left', padx=5
        )
        
        self.status_var = ctk.StringVar(value=self.contact_data.get('status', 'active'))
        status_options = ['active', 'inactive', 'donotcall', 'pending']
        
        self.status_menu = ctk.CTkOptionMenu(
            contact_frame,
            values=status_options,
            variable=self.status_var,
            font=("Arial", 10),
            button_color="#FF6B6B"
        )
        self.status_menu.pack(side='left', padx=5, fill='x', expand=0)
        
        ctk.CTkLabel(contact_frame, text="Teléfono:", font=("Arial", 10)).pack(
            side='left', padx=5
        )
        
        self.phone_entry = ctk.CTkEntry(
            contact_frame,
            placeholder_text="Teléfono",
            font=("Arial", 10),
            height=28
        )
        self.phone_entry.pack(side='left', fill='x', expand=True, padx=5)
        self.phone_entry.insert(0, self.contact_data.get('phone', ''))
        
        # --- Fila 3: Notas (max 244 caracteres) ---
        notes_frame = ctk.CTkFrame(main_frame)
        notes_frame.pack(fill='both', expand=True, pady=5)
        
        ctk.CTkLabel(notes_frame, text="Notas (máx 244 caracteres):", font=("Arial", 10)).pack(
            anchor='w', padx=5, pady=(0, 3)
        )
        
        # Frame para textbox + contador
        notes_input_frame = ctk.CTkFrame(notes_frame)
        notes_input_frame.pack(fill='both', expand=True, padx=5)
        
        self.notes_text = ctk.CTkTextbox(
            notes_input_frame,
            font=("Arial", 10),
            height=80,
            wrap='word'
        )
        self.notes_text.pack(side='left', fill='both', expand=True)
        self.notes_text.insert(1.0, self.contact_data.get('notes', ''))
        self.notes_text.bind('<KeyRelease>', self._on_notes_changed)
        
        self.notes_label = ctk.CTkLabel(
            notes_input_frame,
            text="0/244",
            font=("Arial", 9),
            text_color="gray"
        )
        self.notes_label.pack(side='right', padx=5, pady=5)
        
        # --- Fila 4: Botones de acción ---
        action_frame = ctk.CTkFrame(main_frame)
        action_frame.pack(fill='x', pady=10)
        
        self.call_btn = ctk.CTkButton(
            action_frame,
            text="📞 Llamar",
            font=("Arial", 11, "bold"),
            fg_color="#4CAF50",
            hover_color="#45a049",
            height=35
        )
        self.call_btn.pack(side='left', padx=5, fill='x', expand=True)
        
        self.confirm_btn = ctk.CTkButton(
            action_frame,
            text="✓ Confirmar",
            font=("Arial", 11, "bold"),
            fg_color="#2196F3",
            hover_color="#0b7dda",
            height=35,
            state='disabled'
        )
        self.confirm_btn.pack(side='left', padx=5, fill='x', expand=True)
        
        # Botón eliminar pequeño y rojo
        self.delete_btn = ctk.CTkButton(
            action_frame,
            text="🗑️",
            font=("Arial", 14),
            fg_color="#FF6B6B",
            hover_color="#FF5252",
            width=40,
            height=35
        )
        self.delete_btn.pack(side='left', padx=5)
    
    def _toggle_edit_mode(self):
        """Alternar modo edición"""
        self.is_editing = not self.is_editing
        
        if self.is_editing:
            self.name_entry.configure(state='normal')
            self.edit_btn.configure(text="✓ Guardar", fg_color="#4CAF50")
            self.confirm_btn.configure(state='normal')
        else:
            self.name_entry.configure(state='disabled')
            self.edit_btn.configure(text="✏️ Editar", fg_color="#2196F3")
            self.confirm_btn.configure(state='disabled')
            
            # Guardar datos
            if self.on_save:
                self.contact_data['name'] = self.name_entry.get()
                self.contact_data['status'] = self.status_var.get()
                self.contact_data['phone'] = self.phone_entry.get()
                self.contact_data['notes'] = self.notes_text.get(1.0, 'end-1c')
                self.on_save(self.contact_data)
    
    def _on_notes_changed(self, event=None):
        """Actualizar contador de caracteres en notas"""
        text = self.notes_text.get(1.0, 'end-1c')
        
        if len(text) > 244:
            text = text[:244]
            self.notes_text.delete(1.0, 'end')
            self.notes_text.insert(1.0, text)
        
        self.notes_label.configure(text=f"{len(text)}/244")
    
    def _update_layout(self):
        """Actualizar layout según tamaño de pantalla"""
        # Los widgets se adaptan automáticamente con pack fill/expand
        pass
    
    def get_data(self) -> Dict:
        """Obtener datos del contacto"""
        return {
            'name': self.name_entry.get(),
            'phone': self.phone_entry.get(),
            'status': self.status_var.get(),
            'notes': self.notes_text.get(1.0, 'end-1c')
        }


class ExcelExporter:
    """Exportador a Excel con estilos"""
    
    @staticmethod
    def export_contacts(contacts: List[Dict], filepath: str) -> bool:
        """
        Exportar contactos a Excel
        
        Args:
            contacts: Lista de contactos
            filepath: Ruta del archivo Excel
        
        Returns:
            True si exitoso
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Contactos"
            
            # Estilos
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Headers
            headers = ['Nombre', 'Teléfono', 'Estado', 'Notas', 'Última Llamada', 'Duración']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            
            # Datos
            for contact in contacts:
                ws.append([
                    contact.get('name', ''),
                    contact.get('phone', ''),
                    contact.get('status', ''),
                    contact.get('notes', ''),
                    contact.get('last_call', ''),
                    contact.get('duration', '')
                ])
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 12
            
            # Aplicar bordes y alineación
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            wb.save(filepath)
            logger.info(f"✅ Contactos exportados a: {filepath}")
            return True
        
        except Exception as e:
            logger.error(f"Error exportando a Excel: {e}")
            return False
    
    @staticmethod
    def export_recordings(recordings: List[Dict], filepath: str) -> bool:
        """Exportar grabaciones a Excel"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Grabaciones"
            
            header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            headers = ['ID', 'Contacto', 'Teléfono', 'Agente', 'Inicio', 'Duración (seg)', 'Tamaño (MB)']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            for recording in recordings:
                ws.append([
                    recording.get('recording_id', ''),
                    recording.get('contact_name', ''),
                    recording.get('contact_phone', ''),
                    recording.get('user_name', ''),
                    recording.get('start_time', ''),
                    recording.get('duration_seconds', ''),
                    f"{recording.get('file_size_bytes', 0) / (1024*1024):.2f}"
                ])
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
                for cell in row:
                    cell.border = border
            
            wb.save(filepath)
            logger.info(f"✅ Grabaciones exportadas a: {filepath}")
            return True
        
        except Exception as e:
            logger.error(f"Error exportando grabaciones: {e}")
            return False


class MobileContactsView(ResponsiveFrame):
    """Vista optimizada de contactos para móviles/tablets"""
    
    def __init__(self, master, contacts: List[Dict] = None, **kwargs):
        super().__init__(master, **kwargs)
        self.contacts = contacts or []
        self.selected_contact_idx = None
        
        self._create_widgets()
    
    def _create_widgets(self):
        """Crear interfaz responsiva"""
        # Search bar
        search_frame = ctk.CTkFrame(self)
        search_frame.pack(fill='x', padx=10, pady=10)
        
        self.search_entry = ctk.CTkEntry(
            search_frame,
            placeholder_text="🔍 Buscar contacto...",
            font=("Arial", 11),
            height=35
        )
        self.search_entry.pack(fill='both', expand=True)
        self.search_entry.bind('<KeyRelease>', self._on_search)
        
        # Scrollable frame para contactos
        scroll_frame = ctk.CTkScrollableFrame(self)
        scroll_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        self.contacts_container = scroll_frame
        self._display_contacts()
    
    def _display_contacts(self):
        """Mostrar contactos"""
        # Limpiar
        for widget in self.contacts_container.winfo_children():
            widget.destroy()
        
        # Mostrar cada contacto
        for idx, contact in enumerate(self.contacts):
            self._create_contact_card(idx, contact)
    
    def _create_contact_card(self, idx: int, contact: Dict):
        """Crear tarjeta de contacto"""
        card_frame = ctk.CTkFrame(
            self.contacts_container,
            fg_color="#2b2b2b",
            corner_radius=10
        )
        card_frame.pack(fill='x', pady=8)
        
        # Header de tarjeta
        header = ctk.CTkFrame(card_frame, fg_color="#1f1f1f")
        header.pack(fill='x', padx=10, pady=(10, 5))
        
        ctk.CTkLabel(
            header,
            text=contact.get('name', 'Sin nombre'),
            font=("Arial", 13, "bold")
        ).pack(anchor='w')
        
        ctk.CTkLabel(
            header,
            text=f"📱 {contact.get('phone', 'N/A')}",
            font=("Arial", 10),
            text_color="gray"
        ).pack(anchor='w')
        
        # Estado badge
        status = contact.get('status', 'active')
        status_colors = {
            'active': '#4CAF50',
            'inactive': '#FFC107',
            'donotcall': '#F44336',
            'pending': '#2196F3'
        }
        
        status_badge = ctk.CTkFrame(
            header,
            fg_color=status_colors.get(status, '#666'),
            corner_radius=5
        )
        status_badge.pack(anchor='w', pady=(3, 0))
        
        ctk.CTkLabel(
            status_badge,
            text=f"  {status.upper()}  ",
            font=("Arial", 9, "bold"),
            text_color="white"
        ).pack(padx=5, pady=2)
        
        # Notas
        if contact.get('notes'):
            ctk.CTkLabel(
                card_frame,
                text=f"Notas: {contact.get('notes', '')[:100]}...",
                font=("Arial", 9),
                text_color="#BBBBBB"
            ).pack(anchor='w', padx=15, pady=5)
        
        # Botones
        button_frame = ctk.CTkFrame(card_frame)
        button_frame.pack(fill='x', padx=10, pady=10)
        
        call_btn = ctk.CTkButton(
            button_frame,
            text="📞",
            font=("Arial", 14),
            fg_color="#4CAF50",
            width=45,
            height=35
        )
        call_btn.pack(side='left', padx=3)
        
        edit_btn = ctk.CTkButton(
            button_frame,
            text="✏️",
            font=("Arial", 14),
            fg_color="#2196F3",
            width=45,
            height=35
        )
        edit_btn.pack(side='left', padx=3)
        
        delete_btn = ctk.CTkButton(
            button_frame,
            text="🗑️",
            font=("Arial", 14),
            fg_color="#FF6B6B",
            width=45,
            height=35
        )
        delete_btn.pack(side='left', padx=3)
    
    def _on_search(self, event=None):
        """Filtrar contactos por búsqueda"""
        search_term = self.search_entry.get().lower()
        
        # Filtrar contactos
        filtered = [
            c for c in self.contacts
            if search_term in c.get('name', '').lower() or
               search_term in c.get('phone', '').lower()
        ]
        
        # Limpiar y mostrar
        for widget in self.contacts_container.winfo_children():
            widget.destroy()
        
        for idx, contact in enumerate(filtered):
            self._create_contact_card(idx, contact)


# Atajos de teclado
KEYBOARD_SHORTCUTS = {
    '<Control-n>': 'new_contact',      # Ctrl+N: Nuevo contacto
    '<Control-e>': 'export_excel',     # Ctrl+E: Exportar Excel
    '<Control-f>': 'search',           # Ctrl+F: Buscar
    '<Control-c>': 'call',             # Ctrl+C: Llamar
    '<F2>': 'edit',                    # F2: Editar
    '<Delete>': 'delete_confirm',      # Delete: Eliminar
    '<Escape>': 'cancel',              # Escape: Cancelar
}


def setup_keyboard_shortcuts(root: ctk.CTk, handler_func: Callable):
    """Configurar atajos de teclado"""
    for shortcut, action in KEYBOARD_SHORTCUTS.items():
        root.bind(shortcut, lambda e, a=action: handler_func(a))
    
    logger.info("⌨️ Atajos de teclado configurados")
